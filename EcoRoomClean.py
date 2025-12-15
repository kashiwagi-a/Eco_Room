import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import openpyxl
from datetime import datetime, timedelta
import os
import subprocess
import platform
import re
import csv

# IME関連警告を抑制
if platform.system() == "Darwin":
    os.environ['TK_SILENCE_DEPRECATION'] = '1'


class HotelCleaningSystem:
    def __init__(self):
        self.db_file = "hotel_cleaning.db"
        self.excel_file = "hotel_cleaning_now.xlsx"
        self.records = []
        self.existing_rooms = set()

        # GUI設定
        self.root = tk.Tk()
        self.root.title("客室清掃管理システム")
        self.root.geometry("400x350")

        self.init_database()

        # 起動時にバックアップとチェックアウト削除を実行
        self.startup_cleanup()

        self.setup_gui()
        self.load_data()

    def startup_cleanup(self):
        """起動時のバックアップ作成とチェックアウト削除"""
        # バックアップ作成
        backup_file = self.create_database_backup_silent()

        # チェックアウト削除ダイアログを表示
        checkout_date = self.show_checkout_cleanup_dialog()

        if checkout_date:
            # C/Oの部屋と空白の部屋を削除
            deleted_info = self.cleanup_checkout_rooms(checkout_date)

            if deleted_info['total'] > 0:
                message = f"データ整理が完了しました。\n\n"
                message += f"✓ 削除された部屋: {deleted_info['total']}件\n"
                if deleted_info['checkout'] > 0:
                    message += f"  - チェックアウト完了: {deleted_info['checkout']}件\n"
                if deleted_info['empty'] > 0:
                    message += f"  - 空白データ: {deleted_info['empty']}件\n"
                message += f"✓ チェックアウト日: {checkout_date.strftime('%Y年%m月%d日')}\n"
                message += f"✓ バックアップ: {backup_file}"

                messagebox.showinfo("起動時整理完了", message)

    def show_checkout_cleanup_dialog(self):
        """チェックアウト削除用の日付選択ダイアログ"""
        dialog = tk.Toplevel(self.root)
        dialog.title("起動時データ整理")
        dialog.geometry("450x350")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="20")
        frame.pack(fill="both", expand=True)

        # タイトル
        title_label = ttk.Label(frame, text="起動時データ整理", font=("", 14, "bold"))
        title_label.pack(pady=(0, 15))

        # 説明
        info_label = ttk.Label(frame,
                               text="チェックアウト完了の部屋を削除します。\n"
                                    "指定した日付にC/Oステータスの部屋がデータベースから削除されます。\n"
                                    "（バックアップは自動で作成済みです）",
                               font=("", 10), justify=tk.CENTER)
        info_label.pack(pady=(0, 20))

        # 日付選択
        date_frame = ttk.LabelFrame(frame, text="チェックアウト日", padding="15")
        date_frame.pack(fill="x", pady=(0, 20))

        # デフォルトは今日の日付
        today = datetime.now()
        checkout_vars = {
            'year': tk.StringVar(value=str(today.year)),
            'month': tk.StringVar(value=str(today.month)),
            'day': tk.StringVar(value=str(today.day))
        }

        date_input_frame = ttk.Frame(date_frame)
        date_input_frame.pack()

        ttk.Entry(date_input_frame, textvariable=checkout_vars['year'], width=6).pack(side="left")
        ttk.Label(date_input_frame, text="年").pack(side="left", padx=(0, 10))

        ttk.Entry(date_input_frame, textvariable=checkout_vars['month'], width=4).pack(side="left")
        ttk.Label(date_input_frame, text="月").pack(side="left", padx=(0, 10))

        ttk.Entry(date_input_frame, textvariable=checkout_vars['day'], width=4).pack(side="left")
        ttk.Label(date_input_frame, text="日").pack(side="left")

        # 今日の日付ボタン
        today_frame = ttk.Frame(date_frame)
        today_frame.pack(pady=(10, 0))

        def set_today():
            today = datetime.now()
            checkout_vars['year'].set(str(today.year))
            checkout_vars['month'].set(str(today.month))
            checkout_vars['day'].set(str(today.day))

        ttk.Button(today_frame, text="今日の日付", command=set_today).pack()

        # 結果を格納する変数
        result = {'date': None}

        # ボタン
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill="x")

        def on_ok():
            try:
                checkout_date = datetime(
                    int(checkout_vars['year'].get()),
                    int(checkout_vars['month'].get()),
                    int(checkout_vars['day'].get())
                )
                result['date'] = checkout_date
                dialog.destroy()
            except ValueError:
                messagebox.showerror("エラー", "正しい日付を入力してください")

        def on_skip():
            dialog.destroy()

        ttk.Button(button_frame, text="削除実行", command=on_ok).pack(side="left", padx=5)
        ttk.Button(button_frame, text="スキップ", command=on_skip).pack(side="left", padx=5)

        # ダイアログが閉じられるまで待機
        dialog.wait_window()

        return result['date']

    def cleanup_checkout_rooms(self, checkout_date):
        """指定日より前にC/Oステータスの部屋と空白の部屋を削除"""
        cursor = self.conn.cursor()

        # 指定日以前にC/Oステータスの部屋を検索
        checkout_date_str = checkout_date.strftime('%Y-%m-%d')
        cursor.execute("""
                       SELECT DISTINCT room_number
                       FROM cleaning_schedule
                       WHERE cleaning_date <= ?
                         AND cleaning_status = 'C/O'
                       """, (checkout_date_str,))

        checkout_rooms = [row[0] for row in cursor.fetchall()]

        # デバッグ：C/O部屋を表示
        print(f"C/O部屋: {checkout_rooms}")

        # 清掃スケジュールが空白（全く登録されていない）部屋を検索
        cursor.execute("""
                       SELECT room_number
                       FROM rooms
                       WHERE room_number NOT IN (SELECT DISTINCT room_number
                                                 FROM cleaning_schedule)
                       """)

        empty_rooms = [row[0] for row in cursor.fetchall()]

        # デバッグ：空白部屋を表示
        print(f"空白部屋: {empty_rooms}")

        # さらに、cleaning_statusが空文字列やNULLの部屋も検索
        cursor.execute("""
                       SELECT DISTINCT room_number
                       FROM cleaning_schedule
                       WHERE cleaning_status = ''
                          OR cleaning_status IS NULL
                       """)

        null_or_empty_rooms = [row[0] for row in cursor.fetchall()]
        print(f"空文字列/NULL部屋: {null_or_empty_rooms}")

        # 全ての空白パターンを統合
        rooms_to_delete = list(set(checkout_rooms + empty_rooms + null_or_empty_rooms))

        # デバッグ：削除対象の部屋を表示
        print(f"削除対象の部屋: {rooms_to_delete}")

        if rooms_to_delete:
            # rooms テーブルから削除
            placeholders = ','.join(['?' for _ in rooms_to_delete])
            cursor.execute(f"DELETE FROM rooms WHERE room_number IN ({placeholders})", rooms_to_delete)

            # cleaning_schedule テーブルからも削除
            cursor.execute(f"DELETE FROM cleaning_schedule WHERE room_number IN ({placeholders})", rooms_to_delete)

            self.conn.commit()

            # 削除の内訳を返す
            return {
                'total': len(rooms_to_delete),
                'checkout': len(checkout_rooms),
                'empty': len(empty_rooms),
                'null_or_empty': len(null_or_empty_rooms)
            }

        return {'total': 0, 'checkout': 0, 'empty': 0, 'null_or_empty': 0}

    def create_database_backup_silent(self):
        """サイレントバックアップ作成（メッセージなし）"""
        try:
            import shutil
            backup_name = f"hotel_cleaning_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(self.db_file, backup_name)
            return backup_name
        except Exception as e:
            print(f"バックアップエラー: {e}")
            return "作成失敗"

    def setup_gui(self):
        frame = ttk.Frame(self.root, padding="10")
        frame.pack(fill="both", expand=True)

        # 入力フィールド
        fields = [
            ("部屋番号", "room"),
            ("お客様名", "guest"),
            ("チェックイン日", "date"),
            ("清掃期間（日）", "days")
        ]

        self.vars = {}
        for i, (label, key) in enumerate(fields):
            ttk.Label(frame, text=f"{label}：").grid(row=i, column=0, sticky="w", pady=5)

            if key == "date":
                date_frame = ttk.Frame(frame)
                date_frame.grid(row=i, column=1, sticky="w")
                today = datetime.now()
                self.vars['year'] = tk.StringVar(value=str(today.year))
                self.vars['month'] = tk.StringVar(value=str(today.month))
                self.vars['day'] = tk.StringVar(value=str(today.day))

                for var, label_text in [('year', '年'), ('month', '月'), ('day', '日')]:
                    ttk.Entry(date_frame, textvariable=self.vars[var], width=5).pack(side="left")
                    ttk.Label(date_frame, text=label_text).pack(side="left")
            elif key == "days":
                self.vars[key] = tk.IntVar(value=2)
                ttk.Spinbox(frame, from_=2, to=999, textvariable=self.vars[key], width=20).grid(row=i, column=1,
                                                                                                sticky="w")
            else:
                self.vars[key] = tk.StringVar()
                entry = ttk.Entry(frame, textvariable=self.vars[key], width=20)
                entry.grid(row=i, column=1, sticky="w")
                if key == "room":
                    entry.bind('<KeyRelease>', self.filter_numbers)

        # チェックボックス
        ttk.Label(frame, text="オプション：").grid(row=4, column=0, sticky="w", pady=5)
        self.vars['ecodoor'] = tk.BooleanVar()
        ttk.Checkbutton(frame, text="エコドア", variable=self.vars['ecodoor']).grid(row=4, column=1, sticky="w")

        ttk.Label(frame, text="プラン：").grid(row=5, column=0, sticky="w", pady=5)
        self.vars['ecoplan'] = tk.BooleanVar()
        ttk.Checkbutton(frame, text="エコプラン", variable=self.vars['ecoplan']).grid(row=5, column=1, sticky="w")

        # ボタン（シンプル化）
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=20)

        ttk.Button(btn_frame, text="次の部屋", command=self.add_room).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="エコ票作成", command=self.create_schedule).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="部屋編集", command=self.edit_room).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="CSV読込", command=self.import_csv).pack(side="left", padx=5)

    def filter_numbers(self, event):
        text = self.vars['room'].get()
        filtered = re.sub(r'[^0-9]', '', text)
        if text != filtered:
            self.vars['room'].set(filtered)

    def init_database(self):
        self.conn = sqlite3.connect(self.db_file)
        cursor = self.conn.cursor()

        cursor.execute('''CREATE TABLE IF NOT EXISTS rooms
                          (
                              room_number
                              TEXT
                              PRIMARY
                              KEY,
                              guest_name
                              TEXT,
                              check_in_date
                              DATE,
                              cleaning_days
                              INTEGER,
                              is_ecodoor
                              BOOLEAN,
                              is_ecoplan
                              BOOLEAN
                          )''')

        cursor.execute('''CREATE TABLE IF NOT EXISTS cleaning_schedule
                          (
                              room_number
                              TEXT,
                              cleaning_date
                              DATE,
                              cleaning_status
                              TEXT
                          )''')

        self.conn.commit()

    def load_data(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM rooms ORDER BY CAST(room_number AS INTEGER)")

        for row in cursor.fetchall():
            room, guest, date_str, days, ecodoor, ecoplan = row
            self.existing_rooms.add(room)

            record = {
                'room': room,
                'guest': guest,
                'date': datetime.strptime(date_str, '%Y-%m-%d'),
                'days': days,
                'ecodoor': bool(ecodoor),
                'ecoplan': bool(ecoplan),
                'schedule': {},
                'is_new': False
            }

            # スケジュール読み込み
            cursor.execute(
                "SELECT cleaning_date, cleaning_status FROM cleaning_schedule WHERE room_number = ? ORDER BY cleaning_date",
                (room,))
            for date_str, status in cursor.fetchall():
                date = datetime.strptime(date_str, '%Y-%m-%d')
                record['schedule'][f"{date.month}/{date.day}"] = status

            self.records.append(record)

    def add_room(self):
        room = self.vars['room'].get().strip()
        if not room:
            messagebox.showerror("エラー", "部屋番号を入力してください")
            return

        if room in self.existing_rooms or any(r['room'] == room for r in self.records):
            messagebox.showerror("エラー", "この部屋番号は既に登録されています")
            return

        try:
            date = datetime(int(self.vars['year'].get()), int(self.vars['month'].get()), int(self.vars['day'].get()))
        except ValueError:
            messagebox.showerror("エラー", "正しい日付を入力してください")
            return

        record = {
            'room': room,
            'guest': self.vars['guest'].get().strip(),
            'date': date,
            'days': self.vars['days'].get(),
            'ecodoor': self.vars['ecodoor'].get(),
            'ecoplan': self.vars['ecoplan'].get(),
            'schedule': {}
        }

        # スケジュール生成
        current = date
        for day in range(record['days'] + 1):
            date_str = f"{current.month}/{current.day}"
            if day == 0:
                status = "C/I"
            elif day == record['days']:
                status = "C/O"
            else:
                status = "〇" if day % 3 == 0 else ("エコドア" if record['ecodoor'] else "×")

            record['schedule'][date_str] = status
            current += timedelta(days=1)

        self.records.append(record)
        self.existing_rooms.add(room)
        record['is_new'] = True

        # フォームクリア
        for key in ['room', 'guest']:
            self.vars[key].set("")
        for key in ['ecodoor', 'ecoplan']:
            self.vars[key].set(False)

        messagebox.showinfo("情報", "部屋が追加されました")

    def import_csv(self):
        """CSVファイルを読み込んでエコ清掃対象部屋を選択するダイアログを表示"""
        # ファイル選択ダイアログ
        file_path = filedialog.askopenfilename(
            title="CSVファイルを選択",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")]
        )

        if not file_path:
            return

        try:
            # CSVファイルを読み込み
            eco_rooms = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    if len(row) >= 7:
                        room_number = row[1]  # 2列目：部屋番号
                        room_status = row[6]  # 7列目：部屋の状態

                        # 状態が'1'または'4'の部屋のみ抽出
                        if room_status in ['1', '3']:
                            eco_rooms.append({
                                'room': room_number,
                                'status': room_status
                            })

            if not eco_rooms:
                messagebox.showinfo("情報", "エコ清掃対象の部屋（状態:1または4）が見つかりませんでした。")
                return

            # 部屋選択ダイアログを表示
            self.show_csv_room_selection_dialog(eco_rooms)

        except Exception as e:
            messagebox.showerror("エラー", f"CSVファイルの読み込みに失敗しました: {e}")

    def show_csv_room_selection_dialog(self, eco_rooms):
        """CSVから読み込んだ部屋の選択ダイアログを表示"""
        dialog = tk.Toplevel(self.root)
        dialog.title("エコ清掃部屋選択")
        dialog.geometry("500x700")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="15")
        frame.pack(fill="both", expand=True)

        # タイトル
        title_label = ttk.Label(frame, text="エコ清掃対象部屋の選択", font=("", 14, "bold"))
        title_label.pack(pady=(0, 10))

        # 説明
        info_label = ttk.Label(frame,
                               text=f"CSVから{len(eco_rooms)}件のエコ清掃対象部屋が見つかりました。\n"
                                    "チェックを入れた部屋を2泊宿泊として登録します。",
                               font=("", 10), justify=tk.CENTER)
        info_label.pack(pady=(0, 10))

        # チェックイン日選択
        checkin_frame = ttk.LabelFrame(frame, text="チェックイン日", padding="10")
        checkin_frame.pack(fill="x", pady=(0, 10))

        # デフォルトは明日の日付
        tomorrow = datetime.now() + timedelta(days=1)
        checkin_vars = {
            'year': tk.StringVar(value=str(tomorrow.year)),
            'month': tk.StringVar(value=str(tomorrow.month)),
            'day': tk.StringVar(value=str(tomorrow.day))
        }

        date_input_frame = ttk.Frame(checkin_frame)
        date_input_frame.pack()

        ttk.Entry(date_input_frame, textvariable=checkin_vars['year'], width=6).pack(side="left")
        ttk.Label(date_input_frame, text="年").pack(side="left", padx=(0, 5))

        ttk.Entry(date_input_frame, textvariable=checkin_vars['month'], width=4).pack(side="left")
        ttk.Label(date_input_frame, text="月").pack(side="left", padx=(0, 5))

        ttk.Entry(date_input_frame, textvariable=checkin_vars['day'], width=4).pack(side="left")
        ttk.Label(date_input_frame, text="日").pack(side="left")

        # 日付ショートカットボタン
        shortcut_frame = ttk.Frame(checkin_frame)
        shortcut_frame.pack(pady=(5, 0))

        def set_today():
            today = datetime.now()
            checkin_vars['year'].set(str(today.year))
            checkin_vars['month'].set(str(today.month))
            checkin_vars['day'].set(str(today.day))

        def set_tomorrow():
            tomorrow = datetime.now() + timedelta(days=1)
            checkin_vars['year'].set(str(tomorrow.year))
            checkin_vars['month'].set(str(tomorrow.month))
            checkin_vars['day'].set(str(tomorrow.day))

        ttk.Button(shortcut_frame, text="今日", command=set_today, width=8).pack(side="left", padx=3)
        ttk.Button(shortcut_frame, text="明日", command=set_tomorrow, width=8).pack(side="left", padx=3)

        # 清掃ステータス選択
        status_frame = ttk.LabelFrame(frame, text="中日の清掃ステータス", padding="10")
        status_frame.pack(fill="x", pady=(0, 10))

        cleaning_status_var = tk.StringVar(value="×")  # デフォルトは「×」

        ttk.Radiobutton(status_frame, text="×", variable=cleaning_status_var, value="×").pack(side="left", padx=10)
        ttk.Radiobutton(status_frame, text="エコドア", variable=cleaning_status_var, value="エコドア").pack(side="left",
                                                                                                            padx=10)

        # 全選択/全解除ボタン
        select_frame = ttk.Frame(frame)
        select_frame.pack(fill="x", pady=(0, 10))

        # チェックボックス用の変数を格納
        check_vars = {}

        def select_all():
            for var in check_vars.values():
                var.set(True)

        def deselect_all():
            for var in check_vars.values():
                var.set(False)

        ttk.Button(select_frame, text="全選択", command=select_all).pack(side="left", padx=5)
        ttk.Button(select_frame, text="全解除", command=deselect_all).pack(side="left", padx=5)

        # 選択数表示ラベル
        count_label = ttk.Label(select_frame, text="選択: 0件")
        count_label.pack(side="right", padx=5)

        def update_count(*args):
            selected = sum(1 for var in check_vars.values() if var.get())
            count_label.config(text=f"選択: {selected}件")

        # スクロール可能なフレーム
        list_frame = ttk.LabelFrame(frame, text="部屋一覧", padding="10")
        list_frame.pack(fill="both", expand=True, pady=(0, 15))

        canvas = tk.Canvas(list_frame, width=420, height=350, bg="white")
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # マウスホイールスクロール
        def on_mousewheel(event):
            if platform.system() == "Darwin":
                canvas.yview_scroll(int(-1 * event.delta), "units")
            else:
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind("<MouseWheel>", on_mousewheel)
        scrollable_frame.bind("<MouseWheel>", on_mousewheel)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ヘッダー
        header_frame = ttk.Frame(scrollable_frame)
        header_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(header_frame, text="選択", width=6, font=("", 9, "bold")).pack(side="left", padx=5)
        ttk.Label(header_frame, text="部屋番号", width=10, font=("", 9, "bold")).pack(side="left", padx=5)
        ttk.Label(header_frame, text="状態", width=6, font=("", 9, "bold")).pack(side="left", padx=5)
        ttk.Label(header_frame, text="登録状況", width=15, font=("", 9, "bold")).pack(side="left", padx=5)

        # 部屋ごとのチェックボックス
        for room_info in eco_rooms:
            room_number = room_info['room']
            status = room_info['status']

            row_frame = ttk.Frame(scrollable_frame)
            row_frame.pack(fill="x", pady=2)

            # 既存チェック
            is_existing = room_number in self.existing_rooms or any(r['room'] == room_number for r in self.records)
            status_text = "登録済み" if is_existing else "未登録"
            status_color = "gray" if is_existing else "green"

            check_var = tk.BooleanVar(value=not is_existing)  # 未登録の部屋はデフォルトでチェック
            check_vars[room_number] = check_var
            check_var.trace('w', update_count)

            cb = ttk.Checkbutton(row_frame, variable=check_var)
            cb.pack(side="left", padx=5)
            if is_existing:
                cb.config(state='disabled')

            ttk.Label(row_frame, text=room_number, width=10).pack(side="left", padx=5)
            ttk.Label(row_frame, text=status, width=6).pack(side="left", padx=5)
            status_lbl = ttk.Label(row_frame, text=status_text, width=15)
            status_lbl.pack(side="left", padx=5)

            row_frame.bind("<MouseWheel>", on_mousewheel)

        # 初期カウント更新
        update_count()

        # ボタン
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill="x")

        def register_rooms():
            """選択された部屋を2泊宿泊として登録"""
            selected_rooms = [room for room, var in check_vars.items() if var.get()]

            if not selected_rooms:
                messagebox.showwarning("警告", "部屋が選択されていません。")
                return

            # 選択されたチェックイン日を取得
            try:
                checkin_date = datetime(
                    int(checkin_vars['year'].get()),
                    int(checkin_vars['month'].get()),
                    int(checkin_vars['day'].get())
                )
            except ValueError:
                messagebox.showerror("エラー", "正しいチェックイン日を入力してください。")
                return

            registered_count = 0

            for room_number in selected_rooms:
                # 既存チェック
                if room_number in self.existing_rooms or any(r['room'] == room_number for r in self.records):
                    continue

                # 2泊宿泊として登録
                record = {
                    'room': room_number,
                    'guest': '',  # 空欄
                    'date': checkin_date,
                    'days': 2,  # 2泊
                    'ecodoor': cleaning_status_var.get() == "エコドア",  # 選択に応じて設定
                    'ecoplan': False,
                    'schedule': {},
                    'is_new': True
                }

                # スケジュール生成（2泊）
                current = checkin_date
                for day in range(3):  # 0=C/I, 1=中日, 2=C/O
                    date_str = f"{current.month}/{current.day}"
                    if day == 0:
                        status = "C/I"
                    elif day == 2:
                        status = "C/O"
                    else:
                        status = cleaning_status_var.get()  # 選択したステータスを使用

                    record['schedule'][date_str] = status
                    current += timedelta(days=1)

                self.records.append(record)
                self.existing_rooms.add(room_number)
                registered_count += 1

            dialog.destroy()

            if registered_count > 0:
                messagebox.showinfo("完了",
                                    f"{registered_count}件の部屋を2泊宿泊として登録しました。\n"
                                    f"チェックイン日: {checkin_date.strftime('%Y年%m月%d日')}\n\n"
                                    "「エコ票作成」ボタンでExcelを生成してください。")
            else:
                messagebox.showinfo("情報", "新規登録された部屋はありませんでした。")

        ttk.Button(button_frame, text="選択した部屋を登録", command=register_rooms).pack(side="left", padx=5)
        ttk.Button(button_frame, text="キャンセル", command=dialog.destroy).pack(side="left", padx=5)

    def create_schedule(self):
        """シンプル化されたエコ票作成"""
        if not self.records:
            messagebox.showerror("エラー", "記録する部屋がありません")
            return

        # 入力中の部屋があれば追加
        if self.vars['room'].get().strip():
            self.add_room()

        try:
            # 新しいレコードをデータベース保存
            new_records = [r for r in self.records if r.get('is_new', True)]

            cursor = self.conn.cursor()
            for record in new_records:
                cursor.execute('''INSERT OR REPLACE INTO rooms VALUES (?, ?, ?, ?, ?, ?)''',
                               (record['room'], record['guest'], record['date'].strftime('%Y-%m-%d'),
                                record['days'], record['ecodoor'], record['ecoplan']))

                cursor.execute("DELETE FROM cleaning_schedule WHERE room_number = ?", (record['room'],))
                for date_str, status in record['schedule'].items():
                    month, day = map(int, date_str.split('/'))
                    year = record['date'].year
                    if month < record['date'].month:
                        year += 1
                    cleaning_date = datetime(year, month, day)
                    cursor.execute("INSERT INTO cleaning_schedule VALUES (?, ?, ?)",
                                   (record['room'], cleaning_date.strftime('%Y-%m-%d'), status))

            self.conn.commit()

            # Excel生成
            self.generate_excel()

            # データ再読み込み
            self.records.clear()
            self.existing_rooms.clear()
            self.load_data()

            # 新しいレコードのフラグをクリア
            for record in self.records:
                record['is_new'] = False

            messagebox.showinfo("完了", f"エコ票を作成しました。\n新規登録: {len(new_records)}件")

            # Excel ファイルを開く
            self.open_excel()

        except Exception as e:
            messagebox.showerror("エラー", f"処理中にエラーが発生しました: {e}")

    def generate_excel(self):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ソート
        self.records.sort(key=lambda x: int(x['room']) if x['room'].isdigit() else float('inf'))

        # 全レコードから使用されている月を収集
        months_used = set()
        for record in self.records:
            for date_str in record['schedule'].keys():
                month = int(date_str.split('/')[0])
                months_used.add(month)

        # 月ごとにシートを作成
        for month in sorted(months_used):
            month_name = f"{month}月"
            ws = wb.create_sheet(title=month_name)

            # ヘッダー
            ws.cell(3, 1, "氏名")
            ws.cell(3, 3, "部屋番号")

            # その月の日数を取得
            import calendar
            year = self.records[0]['date'].year if self.records else datetime.now().year
            days_in_month = calendar.monthrange(year, month)[1]

            # 日付ヘッダー
            for day in range(1, days_in_month + 1):
                ws.cell(3, 3 + day, str(day))

            ws.cell(3, 3 + days_in_month + 1, "エコプラン")

            # この月にスケジュールがある部屋のみをフィルタ
            records_this_month = []
            for record in self.records:
                has_schedule_this_month = any(
                    int(date_str.split('/')[0]) == month
                    for date_str in record['schedule'].keys()
                )
                if has_schedule_this_month:
                    records_this_month.append(record)

            # データ
            for i, record in enumerate(records_this_month):
                row = 4 + i
                ws.cell(row, 1, record['guest'])
                ws.cell(row, 3, record['room'])
                ws.cell(row, 3 + days_in_month + 1, "エコプラン" if record['ecoplan'] else "")

                # この月のスケジュールのみ
                for date_str, status in record['schedule'].items():
                    if int(date_str.split('/')[0]) == month:
                        day = int(date_str.split('/')[1])
                        if day <= days_in_month:
                            ws.cell(row, 3 + day, status)

            # 列幅調整
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['C'].width = 8
            for day in range(1, days_in_month + 1):
                col_letter = openpyxl.utils.get_column_letter(3 + day)
                ws.column_dimensions[col_letter].width = 6

        wb.save(self.excel_file)
        wb.close()

    def edit_room(self):
        """部屋の編集"""
        room = self.vars['room'].get().strip()

        if not room:
            self.show_room_edit_dialog()
            return

        self.open_edit_dialog(room)

    def show_room_edit_dialog(self):
        """部屋選択ダイアログを表示（編集用）"""
        all_rooms = list(self.existing_rooms) + [r['room'] for r in self.records]
        all_rooms = sorted(set(all_rooms), key=lambda x: int(x) if x.isdigit() else float('inf'))

        if not all_rooms:
            messagebox.showinfo("情報", "編集できる部屋がありません")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("部屋編集")
        dialog.geometry("250x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text="編集する部屋を選択してください：").pack(pady=10)

        listbox = tk.Listbox(dialog, selectmode=tk.SINGLE)
        for room in all_rooms:
            record = self.find_room_record(room)
            if record:
                display_text = f"{room} - {record['guest']} ({record['days']}日)"
            else:
                display_text = f"{room} - (詳細不明)"
            listbox.insert(tk.END, display_text)
        listbox.pack(fill="both", expand=True, padx=10, pady=5)

        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)

        def on_edit():
            selection = listbox.curselection()
            if selection:
                selected_room = all_rooms[selection[0]]
                dialog.destroy()
                self.open_edit_dialog(selected_room)
            else:
                messagebox.showwarning("警告", "部屋を選択してください。")

        ttk.Button(button_frame, text="編集", command=on_edit).pack(side="left", padx=5)
        ttk.Button(button_frame, text="キャンセル", command=dialog.destroy).pack(side="left", padx=5)

    def find_room_record(self, room_number):
        """部屋番号からレコードを検索"""
        for record in self.records:
            if record['room'] == room_number:
                return record

        # データベースから検索
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM rooms WHERE room_number = ?", (room_number,))
        row = cursor.fetchone()
        if row:
            room, guest, date_str, days, ecodoor, ecoplan = row
            return {
                'room': room,
                'guest': guest,
                'date': datetime.strptime(date_str, '%Y-%m-%d'),
                'days': days,
                'ecodoor': bool(ecodoor),
                'ecoplan': bool(ecoplan)
            }
        return None

    def open_edit_dialog(self, room_number):
        """編集ダイアログを開く"""
        record = self.find_room_record(room_number)
        if not record:
            messagebox.showerror("エラー", f"部屋番号 {room_number} の情報が見つかりません")
            return

        dialog = tk.Toplevel(self.root)
        dialog.title(f"部屋 {room_number} の編集")
        dialog.geometry("500x700")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        frame = ttk.Frame(dialog, padding="10")
        frame.pack(fill="both", expand=True)

        # 編集用変数
        edit_vars = {}

        # 部屋番号（編集不可）
        ttk.Label(frame, text="部屋番号：").grid(row=0, column=0, sticky="w", pady=5)
        ttk.Label(frame, text=room_number, font=("", 10, "bold")).grid(row=0, column=1, sticky="w", pady=5)

        # お客様名
        ttk.Label(frame, text="お客様名：").grid(row=1, column=0, sticky="w", pady=5)
        edit_vars['guest'] = tk.StringVar(value=record['guest'])
        ttk.Entry(frame, textvariable=edit_vars['guest'], width=30).grid(row=1, column=1, sticky="w", pady=5)

        # チェックイン日
        ttk.Label(frame, text="チェックイン日：").grid(row=2, column=0, sticky="w", pady=5)
        checkin_frame = ttk.Frame(frame)
        checkin_frame.grid(row=2, column=1, sticky="w", pady=5)

        edit_vars['checkin_year'] = tk.StringVar(value=str(record['date'].year))
        edit_vars['checkin_month'] = tk.StringVar(value=str(record['date'].month))
        edit_vars['checkin_day'] = tk.StringVar(value=str(record['date'].day))

        ttk.Entry(checkin_frame, textvariable=edit_vars['checkin_year'], width=6).pack(side="left")
        ttk.Label(checkin_frame, text="年").pack(side="left")
        ttk.Entry(checkin_frame, textvariable=edit_vars['checkin_month'], width=4).pack(side="left")
        ttk.Label(checkin_frame, text="月").pack(side="left")
        ttk.Entry(checkin_frame, textvariable=edit_vars['checkin_day'], width=4).pack(side="left")
        ttk.Label(checkin_frame, text="日").pack(side="left")

        # チェックアウト日
        ttk.Label(frame, text="チェックアウト日：").grid(row=3, column=0, sticky="w", pady=5)
        checkout_frame = ttk.Frame(frame)
        checkout_frame.grid(row=3, column=1, sticky="w", pady=5)

        checkout_date = record['date'] + timedelta(days=record['days'])
        edit_vars['checkout_year'] = tk.StringVar(value=str(checkout_date.year))
        edit_vars['checkout_month'] = tk.StringVar(value=str(checkout_date.month))
        edit_vars['checkout_day'] = tk.StringVar(value=str(checkout_date.day))

        ttk.Entry(checkout_frame, textvariable=edit_vars['checkout_year'], width=6).pack(side="left")
        ttk.Label(checkout_frame, text="年").pack(side="left")
        ttk.Entry(checkout_frame, textvariable=edit_vars['checkout_month'], width=4).pack(side="left")
        ttk.Label(checkout_frame, text="月").pack(side="left")
        ttk.Entry(checkout_frame, textvariable=edit_vars['checkout_day'], width=4).pack(side="left")
        ttk.Label(checkout_frame, text="日").pack(side="left")

        # 宿泊日数（自動計算・表示のみ）
        ttk.Label(frame, text="宿泊日数：").grid(row=4, column=0, sticky="w", pady=5)
        days_label = ttk.Label(frame, text=f"{record['days']}日", font=("", 10, "bold"))
        days_label.grid(row=4, column=1, sticky="w", pady=5)

        # エコドア
        ttk.Label(frame, text="オプション：").grid(row=5, column=0, sticky="w", pady=5)
        edit_vars['ecodoor'] = tk.BooleanVar(value=record['ecodoor'])
        ttk.Checkbutton(frame, text="エコドア", variable=edit_vars['ecodoor']).grid(row=5, column=1, sticky="w", pady=5)

        # エコプラン
        ttk.Label(frame, text="プラン：").grid(row=6, column=0, sticky="w", pady=5)
        edit_vars['ecoplan'] = tk.BooleanVar(value=record['ecoplan'])
        ttk.Checkbutton(frame, text="エコプラン", variable=edit_vars['ecoplan']).grid(row=6, column=1, sticky="w",
                                                                                      pady=5)

        # 個別スケジュール編集
        ttk.Label(frame, text="スケジュール編集：").grid(row=7, column=0, sticky="nw", pady=5)
        schedule_frame = ttk.Frame(frame)
        schedule_frame.grid(row=7, column=1, sticky="ew", pady=5)

        # スケジュール編集用のScrollable Frame
        canvas = tk.Canvas(schedule_frame, width=400, height=250, bg="white")
        scrollbar = ttk.Scrollbar(schedule_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # スクロール設定
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def on_mousewheel(event):
            if platform.system() == "Darwin":
                canvas.yview_scroll(int(-1 * event.delta), "units")
            else:  # Windows/Linux
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        # マウスホイールバインド
        def bind_mousewheel(widget):
            if platform.system() == "Darwin":  # mac
                widget.bind("<MouseWheel>", on_mousewheel)
            else:  # Windows/Linux
                widget.bind("<MouseWheel>", on_mousewheel)
                widget.bind("<Button-4>", lambda e: canvas.yview_scroll(-1, "units"))
                widget.bind("<Button-5>", lambda e: canvas.yview_scroll(1, "units"))

        bind_mousewheel(canvas)
        bind_mousewheel(scrollable_frame)

        def on_canvas_click(event):
            canvas.focus_set()

        canvas.bind("<Button-1>", on_canvas_click)
        canvas.bind("<Configure>", configure_scroll_region)
        scrollable_frame.bind("<Configure>", configure_scroll_region)

        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def configure_canvas(event):
            canvas.itemconfig(canvas_window, width=event.width)

        canvas.bind('<Configure>', configure_canvas)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # スケジュール項目の辞書
        schedule_vars = {}

        # スケジュール表示を更新する関数
        def update_schedule_display():
            # 既存のウィジェットを削除
            for widget in scrollable_frame.winfo_children():
                widget.destroy()

            schedule_vars.clear()

            try:
                if not all([edit_vars['checkin_year'].get(), edit_vars['checkin_month'].get(),
                            edit_vars['checkin_day'].get(), edit_vars['checkout_year'].get(),
                            edit_vars['checkout_month'].get(), edit_vars['checkout_day'].get()]):
                    error_label = tk.Label(scrollable_frame, text="日付を入力してください", fg="red")
                    error_label.pack()
                    return

                checkin = datetime(
                    int(edit_vars['checkin_year'].get()),
                    int(edit_vars['checkin_month'].get()),
                    int(edit_vars['checkin_day'].get())
                )
                checkout = datetime(
                    int(edit_vars['checkout_year'].get()),
                    int(edit_vars['checkout_month'].get()),
                    int(edit_vars['checkout_day'].get())
                )

                if checkout <= checkin:
                    error_label = tk.Label(scrollable_frame, text="無効な日付範囲", fg="red")
                    error_label.pack()
                    return

                days = (checkout - checkin).days
                days_label.config(text=f"{days}日")

                # 現在のスケジュールを取得
                current_schedule = self.get_room_schedule(room_number)

                # メインコンテナフレーム
                main_container = ttk.Frame(scrollable_frame)
                main_container.pack(fill="both", expand=True, padx=5, pady=5)

                # ヘッダー
                header_frame = ttk.Frame(main_container)
                header_frame.pack(fill="x", pady=(0, 5))

                ttk.Label(header_frame, text="日付", font=("", 9, "bold")).grid(row=0, column=0, padx=5, pady=2,
                                                                                sticky="w")
                ttk.Label(header_frame, text="現在", font=("", 9, "bold")).grid(row=0, column=1, padx=5, pady=2,
                                                                                sticky="w")
                ttk.Label(header_frame, text="変更", font=("", 9, "bold")).grid(row=0, column=2, padx=5, pady=2,
                                                                                sticky="w")

                # データフレーム
                data_frame = ttk.Frame(main_container)
                data_frame.pack(fill="both", expand=True)

                # 日付範囲でスケジュール表示
                current_date = checkin
                row = 0
                while current_date <= checkout:
                    date_str = f"{current_date.month}/{current_date.day}"

                    # 行フレームを作成
                    row_frame = ttk.Frame(data_frame)
                    row_frame.pack(fill="x", pady=1)

                    # 日付表示
                    date_label = ttk.Label(row_frame, text=date_str, width=8)
                    date_label.grid(row=0, column=0, padx=5, pady=2, sticky="w")

                    # 現在の状態を取得
                    if current_date == checkin:
                        current_status = "C/I"
                    elif current_date == checkout:
                        current_status = "C/O"
                    else:
                        current_status = current_schedule.get(date_str, "×")

                    # 現在の状態表示
                    status_label = ttk.Label(row_frame, text=current_status, width=8)
                    status_label.grid(row=0, column=1, padx=5, pady=2, sticky="w")

                    # 変更用コンボボックス
                    schedule_vars[date_str] = tk.StringVar(value=current_status)
                    status_combo = ttk.Combobox(row_frame, textvariable=schedule_vars[date_str], width=8)
                    status_combo['values'] = ('C/I', 'C/O', '〇', '×', 'エコドア')
                    status_combo.grid(row=0, column=2, padx=5, pady=2, sticky="w")

                    # C/IとC/Oは固定（変更不可）
                    if current_date == checkin or current_date == checkout:
                        status_combo.config(state='disabled')

                    # マウスホイールをバインド
                    bind_mousewheel(row_frame)
                    bind_mousewheel(date_label)
                    bind_mousewheel(status_label)

                    current_date += timedelta(days=1)
                    row += 1

                # 全てのウィジェットにマウスホイールをバインド
                bind_mousewheel(main_container)
                bind_mousewheel(header_frame)
                bind_mousewheel(data_frame)

                # スクロール範囲を更新
                def update_scroll_region():
                    scrollable_frame.update_idletasks()
                    canvas.update_idletasks()
                    canvas.configure(scrollregion=canvas.bbox("all"))
                    canvas.yview_moveto(0)

                canvas.after(100, update_scroll_region)

                # 長期間の場合はスクロールバーを目立たせる
                if days > 10:
                    info_label = tk.Label(main_container,
                                          text=f"※{days}日間の長期滞在 - マウスホイールまたはスクロールバーで全日程を確認してください",
                                          fg="blue", font=("", 8), wraplength=350)
                    info_label.pack(pady=5)
                    bind_mousewheel(info_label)

            except ValueError:
                days_label.config(text="日付エラー")
                error_label = tk.Label(scrollable_frame, text="正しい日付を入力してください", fg="red")
                error_label.pack()

        # 日数自動計算機能
        def update_days():
            update_schedule_display()

        # 日付変更時の自動更新
        for var in ['checkin_year', 'checkin_month', 'checkin_day', 'checkout_year', 'checkout_month', 'checkout_day']:
            edit_vars[var].trace('w', lambda *args: update_days())

        # 初期表示
        update_schedule_display()

        # ボタン
        button_frame = ttk.Frame(frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=20)

        def save_changes():
            try:
                checkin_date = datetime(
                    int(edit_vars['checkin_year'].get()),
                    int(edit_vars['checkin_month'].get()),
                    int(edit_vars['checkin_day'].get())
                )
                checkout_date = datetime(
                    int(edit_vars['checkout_year'].get()),
                    int(edit_vars['checkout_month'].get()),
                    int(edit_vars['checkout_day'].get())
                )

                if checkout_date <= checkin_date:
                    messagebox.showerror("エラー", "チェックアウト日はチェックイン日より後である必要があります")
                    return

                calculated_days = (checkout_date - checkin_date).days

            except ValueError:
                messagebox.showerror("エラー", "正しい日付を入力してください")
                return

            # 更新されたレコードを作成
            updated_record = {
                'room': room_number,
                'guest': edit_vars['guest'].get().strip(),
                'date': checkin_date,
                'days': calculated_days,
                'ecodoor': edit_vars['ecodoor'].get(),
                'ecoplan': edit_vars['ecoplan'].get(),
                'schedule': {},
                'is_new': True
            }

            # 個別に編集されたスケジュールを使用
            for date_str, var in schedule_vars.items():
                updated_record['schedule'][date_str] = var.get()

            # メモリ上のレコードを更新
            for i, rec in enumerate(self.records):
                if rec['room'] == room_number:
                    self.records[i] = updated_record
                    break
            else:
                self.records.append(updated_record)

            # データベースに保存
            cursor = self.conn.cursor()
            cursor.execute('''INSERT OR REPLACE INTO rooms VALUES (?, ?, ?, ?, ?, ?)''',
                           (updated_record['room'], updated_record['guest'],
                            updated_record['date'].strftime('%Y-%m-%d'),
                            updated_record['days'], updated_record['ecodoor'], updated_record['ecoplan']))

            cursor.execute("DELETE FROM cleaning_schedule WHERE room_number = ?", (room_number,))
            for date_str, status in updated_record['schedule'].items():
                month, day = map(int, date_str.split('/'))
                year = updated_record['date'].year
                if month < updated_record['date'].month:
                    year += 1
                cleaning_date = datetime(year, month, day)
                cursor.execute("INSERT INTO cleaning_schedule VALUES (?, ?, ?)",
                               (room_number, cleaning_date.strftime('%Y-%m-%d'), status))

            self.conn.commit()

            dialog.destroy()
            messagebox.showinfo("成功", f"部屋 {room_number} の情報を更新しました")

        ttk.Button(button_frame, text="保存", command=save_changes).pack(side="left", padx=5)
        ttk.Button(button_frame, text="キャンセル", command=dialog.destroy).pack(side="left", padx=5)

    def get_room_schedule(self, room_number):
        """部屋のスケジュールを取得"""
        # メモリから検索
        for record in self.records:
            if record['room'] == room_number:
                return record.get('schedule', {})

        # データベースから検索
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT cleaning_date, cleaning_status FROM cleaning_schedule WHERE room_number = ? ORDER BY cleaning_date",
            (room_number,))
        schedule = {}
        for date_str, status in cursor.fetchall():
            date = datetime.strptime(date_str, '%Y-%m-%d')
            schedule[f"{date.month}/{date.day}"] = status

        return schedule

    def open_excel(self):
        """エクセルファイルを開く"""
        try:
            if platform.system() == "Darwin":
                subprocess.call(("open", self.excel_file))
            elif platform.system() == "Windows":
                os.startfile(self.excel_file)
            else:
                subprocess.call(("xdg-open", self.excel_file))
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルを開けませんでした: {e}")

    def run(self):
        """アプリケーションの実行"""
        self.root.protocol("WM_DELETE_WINDOW", lambda: (self.conn.close(), self.root.destroy()))
        self.root.mainloop()


if __name__ == "__main__":
    app = HotelCleaningSystem()
    app.run()